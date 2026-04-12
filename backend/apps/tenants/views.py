from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import Tenant
from .serializers import (
    TenantSerializer, TenantCreateSerializer, TenantUpdateSerializer,
    TenantUserSerializer, UserCreateSerializer
)
from .permissions import IsPlatformAdmin, IsTenantAdmin
from apps.user.models import User, UserRole


class TenantViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing tenants.
    
    List/Retrieve: Platform admins see all, others see only their tenant.
    Create/Update/Delete: Platform admin only.
    """
    serializer_class = TenantSerializer
    permission_classes = [permissions.IsAuthenticated]
    lookup_field = "slug"

    def get_queryset(self):
        user = self.request.user
        # Platform admins (superusers) can see all tenants
        if user.is_superuser:
            return Tenant.objects.all()
        elif user.tenant:
            return Tenant.objects.filter(id=user.tenant.id)
        return Tenant.objects.none()

    def get_serializer_class(self):
        if self.action == "create":
            return TenantCreateSerializer
        if self.action in ["update", "partial_update"]:
            return TenantUpdateSerializer
        return TenantSerializer

    def get_permissions(self):
        if self.action in ["create", "update", "partial_update", "destroy", 
                          "users", "assign_user", "remove_user", "create_user",
                          "bulk_import_tenants", "bulk_import_users", 
                          "export_tenants_template", "export_users_template",
                          "export_tenants", "export_users",
                          "all_users", "create_user_unassigned"]:
            return [IsPlatformAdmin()]
        if self.action in ["my_users", "update_user_role"]:
            return [IsTenantAdmin()]
        return super().get_permissions()

    @action(detail=False, methods=["get"])
    def current(self, request):
        """Get the current user's tenant."""
        if not request.user.tenant:
            return Response(
                {"detail": "You are not assigned to a tenant."},
                status=status.HTTP_404_NOT_FOUND
            )
        serializer = self.get_serializer(request.user.tenant)
        return Response(serializer.data)

    @action(detail=True, methods=["get"])
    def users(self, request, slug=None):
        """Get users assigned to this tenant."""
        tenant = self.get_object()
        users = User.objects.filter(tenant=tenant)
        serializer = TenantUserSerializer(users, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["post"], url_path="assign-user")
    def assign_user(self, request, slug=None):
        """Assign a user to this tenant."""
        tenant = self.get_object()
        user_id = request.data.get("user_id")
        
        if not user_id:
            return Response(
                {"detail": "user_id is required."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response(
                {"detail": "User not found."},
                status=status.HTTP_404_NOT_FOUND
            )
        
        user.tenant = tenant
        user.save()
        return Response({"detail": f"User {user.username} assigned to {tenant.name}."})

    @action(detail=True, methods=["post"], url_path="remove-user")
    def remove_user(self, request, slug=None):
        """Remove a user from this tenant."""
        tenant = self.get_object()
        user_id = request.data.get("user_id")
        
        if not user_id:
            return Response(
                {"detail": "user_id is required."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            user = User.objects.get(id=user_id, tenant=tenant)
        except User.DoesNotExist:
            return Response(
                {"detail": "User not found in this tenant."},
                status=status.HTTP_404_NOT_FOUND
            )
        
        user.tenant = None
        update_fields = ["tenant"]
        if user.role == UserRole.TENANT_ADMIN:
            user.role = UserRole.TENANT_USER
            update_fields.append("role")
        user.save(update_fields=update_fields)
        return Response({"detail": f"User {user.username} removed from {tenant.name}."})

    @action(detail=False, methods=["get"], url_path="available-users")
    def available_users(self, request):
        """Get users that are not assigned to any tenant (for assignment)."""
        users = User.objects.filter(tenant__isnull=True, is_superuser=False)
        serializer = TenantUserSerializer(users, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["post"], url_path="create-user")
    def create_user(self, request, slug=None):
        """Create a new user and assign to this tenant."""
        tenant = self.get_object()
        serializer = UserCreateSerializer(data=request.data)
        
        if serializer.is_valid():
            user = serializer.save()
            user.tenant = tenant
            user.save()
            return Response(
                TenantUserSerializer(user).data,
                status=status.HTTP_201_CREATED
            )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["post"], url_path="create-user-unassigned")
    def create_user_unassigned(self, request):
        """Create a new user without tenant assignment."""
        raw_role = request.data.get("role")
        if raw_role in (None, ""):
            role = UserRole.TENANT_USER
        else:
            role = str(raw_role).strip()
        if role == UserRole.TENANT_ADMIN:
            return Response(
                {
                    "detail": (
                        "A shipyard administrator must belong to a shipyard. "
                        "Create the user in a shipyard context or assign a shipyard first."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        serializer = UserCreateSerializer(data=request.data)

        if serializer.is_valid():
            user = serializer.save()
            return Response(
                TenantUserSerializer(user).data,
                status=status.HTTP_201_CREATED
            )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    # ============ BULK IMPORT/EXPORT ============

    @action(detail=False, methods=["get"], url_path="export-tenants-template")
    def export_tenants_template(self, request):
        """Download Excel template for bulk tenant import."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Shipyards"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
        info_fill = PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid")
        example_fill = PatternFill(start_color="E9D5FF", end_color="E9D5FF", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Headers (row 1); guidance (row 2); example (row 3). Data starts row 4.
        headers = ["Name *", "Description", "Address", "Phone", "Email"]
        info_row = [
            "Shipyard name (required)",
            "Short description",
            "Full address",
            "Phone number",
            "Contact email",
        ]
        example_row = [
            "Example Shipyard",
            "Marine repair yard",
            "Tuzla, Istanbul",
            "+90 216 000 0000",
            "info@example.com",
        ]
        
        for col_idx, (header, info, example) in enumerate(zip(headers, info_row, example_row), 1):
            # Header
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
            # Info
            cell = ws.cell(row=2, column=col_idx, value=info)
            cell.fill = info_fill
            cell.border = thin_border
            
            # Example
            cell = ws.cell(row=3, column=col_idx, value=example)
            cell.fill = example_fill
            cell.border = thin_border
            
            ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 5, 20)
        
        ws.freeze_panes = 'A4'
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="shipyards_template.xlsx"'
        return response

    @action(detail=False, methods=["get"], url_path="export-users-template")
    def export_users_template(self, request):
        """Download Excel template for bulk user import."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
        info_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        example_fill = PatternFill(start_color="A7F3D0", end_color="A7F3D0", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        headers = [
            "Username *",
            "Email *",
            "First name",
            "Last name",
            "Password",
            "Role",
            "Shipyard name",
        ]
        info_row = [
            "Unique username",
            "Valid email",
            "First name",
            "Last name",
            "Min 8 characters (empty defaults to Temp1234!)",
            "tenant_user / tenant_admin (tenant_admin requires shipyard)",
            "Exact shipyard name (empty = unassigned)",
        ]
        example_row = [
            "jdoe",
            "jdoe@example.com",
            "Jane",
            "Doe",
            "Password1!",
            "tenant_user",
            "Example Shipyard",
        ]
        
        for col_idx, (header, info, example) in enumerate(zip(headers, info_row, example_row), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
            cell = ws.cell(row=2, column=col_idx, value=info)
            cell.fill = info_fill
            cell.border = thin_border
            
            cell = ws.cell(row=3, column=col_idx, value=example)
            cell.fill = example_fill
            cell.border = thin_border
            
            ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 5, 25)
        
        ws.freeze_panes = 'A4'
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="users_template.xlsx"'
        return response

    @action(detail=False, methods=["post"], url_path="bulk-import-tenants")
    def bulk_import_tenants(self, request):
        """Import tenants from Excel file."""
        file = request.FILES.get('file')
        if not file:
            return Response({"detail": "File is required."}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            
            results = {"success": 0, "errors": []}
            
            for row_idx in range(4, ws.max_row + 1):
                name = ws.cell(row=row_idx, column=1).value
                if not name or str(name).strip() == '':
                    continue
                
                name = str(name).strip()
                
                # Check if exists
                if Tenant.objects.filter(name=name).exists():
                    results["errors"].append(
                        {"row": row_idx, "errors": [f"A shipyard named '{name}' already exists."]}
                    )
                    continue
                
                try:
                    tenant = Tenant.objects.create(
                        name=name,
                        description=str(ws.cell(row=row_idx, column=2).value or '').strip(),
                        address=str(ws.cell(row=row_idx, column=3).value or '').strip(),
                        phone=str(ws.cell(row=row_idx, column=4).value or '').strip(),
                        email=str(ws.cell(row=row_idx, column=5).value or '').strip(),
                    )
                    results["success"] += 1
                except Exception as e:
                    results["errors"].append({"row": row_idx, "errors": [str(e)]})
            
            return Response({
                "message": f"Imported {results['success']} shipyard(s).",
                "success_count": results["success"],
                "error_count": len(results["errors"]),
                "errors": results["errors"]
            })
        except Exception as e:
            return Response(
                {"detail": f"Error while processing file: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=False, methods=["post"], url_path="bulk-import-users")
    def bulk_import_users(self, request):
        """Import users from Excel file."""
        file = request.FILES.get('file')
        if not file:
            return Response({"detail": "File is required."}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            
            results = {"success": 0, "errors": []}
            
            for row_idx in range(4, ws.max_row + 1):
                username = ws.cell(row=row_idx, column=1).value
                email = ws.cell(row=row_idx, column=2).value
                
                if not username or str(username).strip() == '':
                    continue
                
                username = str(username).strip()
                email = str(email or '').strip()
                
                row_errors = []
                
                if not email:
                    row_errors.append("Email is required.")
                
                if User.objects.filter(username=username).exists():
                    row_errors.append(f"Username '{username}' is already taken.")
                
                if email and User.objects.filter(email=email).exists():
                    row_errors.append(f"Email '{email}' is already in use.")
                
                if row_errors:
                    results["errors"].append({"row": row_idx, "errors": row_errors})
                    continue
                
                first_name = str(ws.cell(row=row_idx, column=3).value or '').strip()
                last_name = str(ws.cell(row=row_idx, column=4).value or '').strip()
                password = str(ws.cell(row=row_idx, column=5).value or 'Temp1234!').strip()
                role = str(ws.cell(row=row_idx, column=6).value or 'tenant_user').strip().lower()
                tenant_name = str(ws.cell(row=row_idx, column=7).value or '').strip()
                
                # Validate role
                if role not in ['tenant_user', 'tenant_admin', 'platform_admin']:
                    role = 'tenant_user'
                
                # Find tenant
                tenant = None
                if tenant_name:
                    tenant = Tenant.objects.filter(name=tenant_name).first()
                    if not tenant:
                        results["errors"].append(
                            {
                                "row": row_idx,
                                "errors": [f"Shipyard '{tenant_name}' was not found."],
                            }
                        )
                        continue

                if role == "tenant_admin" and tenant is None:
                    results["errors"].append(
                        {
                            "row": row_idx,
                            "errors": [
                                "Shipyard name is required when role is tenant_admin."
                            ],
                        }
                    )
                    continue

                try:
                    user = User.objects.create_user(
                        username=username,
                        email=email,
                        password=password,
                        first_name=first_name,
                        last_name=last_name,
                        role=role,
                        tenant=tenant
                    )
                    results["success"] += 1
                except Exception as e:
                    results["errors"].append({"row": row_idx, "errors": [str(e)]})
            
            return Response({
                "message": f"Imported {results['success']} user(s).",
                "success_count": results["success"],
                "error_count": len(results["errors"]),
                "errors": results["errors"]
            })
        except Exception as e:
            return Response(
                {"detail": f"Error while processing file: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=False, methods=["get"], url_path="export-tenants")
    def export_tenants(self, request):
        """Export all tenants to Excel."""
        tenants = Tenant.objects.all()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Shipyards"
        
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
        
        headers = [
            "ID",
            "Name",
            "Slug",
            "Description",
            "Address",
            "Phone",
            "Email",
            "User count",
            "Active",
            "Created",
        ]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 5, 15)
        
        for row_idx, tenant in enumerate(tenants, 2):
            ws.cell(row=row_idx, column=1, value=tenant.id)
            ws.cell(row=row_idx, column=2, value=tenant.name)
            ws.cell(row=row_idx, column=3, value=tenant.slug)
            ws.cell(row=row_idx, column=4, value=tenant.description)
            ws.cell(row=row_idx, column=5, value=tenant.address)
            ws.cell(row=row_idx, column=6, value=tenant.phone)
            ws.cell(row=row_idx, column=7, value=tenant.email)
            ws.cell(row=row_idx, column=8, value=tenant.users.count())
            ws.cell(row=row_idx, column=9, value="Yes" if tenant.is_active else "No")
            ws.cell(row=row_idx, column=10, value=tenant.created_at.strftime('%Y-%m-%d'))
        
        ws.freeze_panes = 'A2'
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="shipyards_export.xlsx"'
        return response

    @action(detail=False, methods=["get"], url_path="all-users")
    def all_users(self, request):
        """Get all users (platform admin only)."""
        users = User.objects.select_related('tenant').all().order_by('-date_joined')
        serializer = TenantUserSerializer(users, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["get"], url_path="my-users")
    def my_users(self, request):
        """Get users of the current user's tenant (tenant admin or platform admin)."""
        if not request.user.tenant:
            return Response(
                {"detail": "You are not assigned to a tenant."},
                status=status.HTTP_404_NOT_FOUND
            )
        users = User.objects.filter(tenant=request.user.tenant).order_by('-date_joined')
        serializer = TenantUserSerializer(users, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["patch"], url_path="update-user-role/(?P<user_id>[^/.]+)")
    def update_user_role(self, request, user_id=None):
        """Update a user's role. Platform admin can update any user, tenant admin can update users in their tenant."""
        try:
            target_user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response({"detail": "User not found."}, status=status.HTTP_404_NOT_FOUND)

        if not request.user.is_superuser:
            if not request.user.is_tenant_admin or target_user.tenant != request.user.tenant:
                return Response(
                    {"detail": "You can only update users in your tenant."},
                    status=status.HTTP_403_FORBIDDEN
                )
            if request.data.get('role') == 'platform_admin':
                return Response(
                    {"detail": "You cannot assign platform admin role."},
                    status=status.HTTP_403_FORBIDDEN
                )

        new_role = request.data.get("role")
        if new_role and new_role in ["platform_admin", "tenant_admin", "tenant_user"]:
            if new_role == UserRole.TENANT_ADMIN and target_user.tenant is None:
                return Response(
                    {
                        "detail": (
                            "Assign this user to a shipyard before making them a shipyard administrator."
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            target_user.role = new_role
            target_user.save(update_fields=["role"])
            return Response(TenantUserSerializer(target_user).data)
        return Response({"detail": "Invalid role."}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["get"], url_path="export-users")
    def export_users(self, request):
        """Export all users to Excel."""
        users = User.objects.select_related('tenant').all()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users"
        
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
        
        headers = [
            "ID",
            "Username",
            "Email",
            "First name",
            "Last name",
            "Role",
            "Shipyard",
            "Active",
            "Last login",
            "Registered",
        ]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 5, 18)
        
        role_labels = {
            "platform_admin": "Platform admin (role)",
            "tenant_admin": "Shipyard administrator",
            "tenant_user": "User",
        }
        
        for row_idx, user in enumerate(users, 2):
            ws.cell(row=row_idx, column=1, value=user.id)
            ws.cell(row=row_idx, column=2, value=user.username)
            ws.cell(row=row_idx, column=3, value=user.email)
            ws.cell(row=row_idx, column=4, value=user.first_name)
            ws.cell(row=row_idx, column=5, value=user.last_name)
            # Show proper role label, check superuser first
            if user.is_superuser:
                role_display = "Superuser"
            else:
                role_display = role_labels.get(user.role, user.role)
            ws.cell(row=row_idx, column=6, value=role_display)
            ws.cell(row=row_idx, column=7, value=user.tenant.name if user.tenant else "-")
            ws.cell(row=row_idx, column=8, value="Yes" if user.is_active else "No")
            ws.cell(row=row_idx, column=9, value=user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else "-")
            ws.cell(row=row_idx, column=10, value=user.date_joined.strftime('%Y-%m-%d'))
        
        ws.freeze_panes = 'A2'
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="users_export.xlsx"'
        return response

