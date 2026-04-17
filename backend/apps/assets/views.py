from rest_framework import viewsets, permissions, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse
from .models import AssetType, Asset
from .serializers import AssetTypeSerializer, AssetSerializer
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import json


class AssetTypeViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing Asset Types (Schemas).
    AssetTypes are global - any authenticated user can view them.
    Only superusers/admins can create/edit/delete.
    """
    serializer_class = AssetTypeSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'description']
    ordering_fields = ['name', 'created_at']

    def get_queryset(self):
        # All authenticated users can see all asset types (they're global)
        return AssetType.objects.all()

    def create(self, request, *args, **kwargs):
        # Only superusers can create global asset types
        if not request.user.is_superuser and not request.user.is_platform_admin:
            return Response(
                {"detail": "Only platform administrators can create asset types."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        if not request.user.is_superuser and not request.user.is_platform_admin:
            return Response(
                {"detail": "Only platform administrators can update asset types."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        if not request.user.is_superuser and not request.user.is_platform_admin:
            return Response(
                {"detail": "Only platform administrators can delete asset types."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['get'])
    def export_template(self, request, pk=None):
        """
        Generate Excel template for bulk asset import.
        Uses field labels (not keys) for user-friendly headers.
        """
        asset_type = self.get_object()
        schema = asset_type.schema or []
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = asset_type.name[:31]  # Excel limit
        
        # Header styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Info row styles
        info_font = Font(italic=True, color="666666", size=10)
        info_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        
        # Build headers and info from schema
        headers = []
        info_row = []
        field_mapping = {}  # Store label -> key mapping in hidden sheet
        
        for field in schema:
            label = field.get('label', field.get('key', 'Alan'))
            key = field.get('key', '')
            field_type = field.get('type', 'text')
            required = field.get('required', False)
            options = field.get('options', [])
            
            # Add required marker
            if required:
                label = f"{label} *"
            
            headers.append(label)
            field_mapping[label.replace(' *', '')] = key
            
            # Build info text
            type_names = {
                'text': 'Metin',
                'number': 'Sayı',
                'date': 'Tarih (YYYY-AA-GG)',
                'select': f"Seçenek: {', '.join(options)}" if options else 'Seçenek',
                'checkbox': 'Evet/Hayır',
                'email': 'E-posta',
                'phone': 'Telefon'
            }
            info_text = type_names.get(field_type, field_type)
            if required:
                info_text = f"[Zorunlu] {info_text}"
            info_row.append(info_text)
        
        # Write headers (Row 1)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            # Auto-adjust column width
            ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 5, 15)
        
        # Write info row (Row 2)
        for col_idx, info in enumerate(info_row, 1):
            cell = ws.cell(row=2, column=col_idx, value=info)
            cell.font = info_font
            cell.fill = info_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        
        # Add example row (Row 3)
        example_fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
        for col_idx, field in enumerate(schema, 1):
            field_type = field.get('type', 'text')
            options = field.get('options', [])
            
            examples = {
                'text': 'Örnek metin',
                'number': '100',
                'date': '2024-01-15',
                'select': options[0] if options else 'Seçenek1',
                'checkbox': 'Evet',
                'email': 'ornek@email.com',
                'phone': '+90 555 123 45 67'
            }
            cell = ws.cell(row=3, column=col_idx, value=examples.get(field_type, ''))
            cell.fill = example_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Store field mapping in hidden sheet for import reference
        mapping_ws = wb.create_sheet("_mapping")
        mapping_ws.sheet_state = 'hidden'
        mapping_ws.cell(row=1, column=1, value=json.dumps(field_mapping, ensure_ascii=False))
        mapping_ws.cell(row=2, column=1, value=str(asset_type.id))
        
        # Freeze header rows
        ws.freeze_panes = 'A4'
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Create response
        filename = f"{asset_type.name.replace(' ', '_')}_sablonu.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response


class AssetViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing Assets.
    Assets are tenant-specific - users only see their tenant's assets.
    """
    serializer_class = AssetSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['asset_type__name']
    ordering_fields = ['created_at', 'updated_at']

    def get_queryset(self):
        user = self.request.user
        
        # Superusers can see all assets (with optional tenant filter)
        if user.is_superuser or user.is_platform_admin:
            queryset = Asset.objects.all()
            tenant_id = self.request.query_params.get('tenant')
            if tenant_id:
                queryset = queryset.filter(tenant_id=tenant_id)
        else:
            # Regular users see only their tenant's assets
            if not user.tenant:
                return Asset.objects.none()
            queryset = Asset.objects.filter(tenant=user.tenant)
        
        # Filter by asset_type
        asset_type_id = self.request.query_params.get('asset_type')
        if asset_type_id:
            queryset = queryset.filter(asset_type_id=asset_type_id)
        
        # Filter by assigned_to
        assigned_to_id = self.request.query_params.get('assigned_to')
        if assigned_to_id:
            queryset = queryset.filter(assigned_to_id=assigned_to_id)
            
        return queryset.select_related('tenant', 'asset_type', 'assigned_to')

    @action(detail=False, methods=['post'])
    def bulk_import(self, request):
        """
        Import assets from Excel file.
        Expects multipart form with 'file' and 'asset_type' fields.
        """
        file = request.FILES.get('file')
        asset_type_id = request.data.get('asset_type')
        tenant_id = request.data.get('tenant')
        
        if not file:
            return Response({"detail": "Dosya gerekli."}, status=status.HTTP_400_BAD_REQUEST)
        
        if not asset_type_id:
            return Response({"detail": "Varlık tipi gerekli."}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            asset_type = AssetType.objects.get(id=asset_type_id)
        except AssetType.DoesNotExist:
            return Response({"detail": "Varlık tipi bulunamadı."}, status=status.HTTP_404_NOT_FOUND)
        
        # Determine tenant
        user = request.user
        if user.is_superuser or user.is_platform_admin:
            if tenant_id:
                from apps.tenants.models import Tenant
                try:
                    tenant = Tenant.objects.get(id=tenant_id)
                except Tenant.DoesNotExist:
                    return Response({"detail": "Tersane bulunamadı."}, status=status.HTTP_404_NOT_FOUND)
            else:
                return Response({"detail": "Superuser için tersane seçimi gerekli."}, status=status.HTTP_400_BAD_REQUEST)
        else:
            tenant = user.tenant
            if not tenant:
                return Response({"detail": "Kullanıcının tersanesi yok."}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            
            # Get schema and build label->key mapping
            schema = asset_type.schema or []
            label_to_key = {}
            field_types = {}
            required_fields = []
            
            for field in schema:
                label = field.get('label', field.get('key', ''))
                key = field.get('key', '')
                label_to_key[label] = key
                label_to_key[f"{label} *"] = key  # Handle required marker
                field_types[key] = field.get('type', 'text')
                if field.get('required'):
                    required_fields.append(key)
            
            # Read headers from row 1
            headers = []
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col_idx).value
                if cell_value:
                    headers.append(str(cell_value).strip())
            
            # Map headers to field keys
            header_keys = []
            for header in headers:
                clean_header = header.replace(' *', '').strip()
                key = label_to_key.get(clean_header) or label_to_key.get(header)
                header_keys.append(key)
            
            # Process data rows (skip header and info rows, start from row 4)
            results = {"success": 0, "errors": []}
            
            for row_idx in range(4, ws.max_row + 1):
                # Check if row is empty
                row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, len(headers) + 1)]
                if all(v is None or str(v).strip() == '' for v in row_values):
                    continue
                
                custom_data = {}
                row_errors = []
                
                for col_idx, key in enumerate(header_keys):
                    if not key:
                        continue
                    
                    cell_value = ws.cell(row=row_idx, column=col_idx + 1).value
                    field_type = field_types.get(key, 'text')
                    
                    # Convert value based on type
                    if cell_value is not None:
                        if field_type == 'checkbox':
                            val_str = str(cell_value).lower().strip()
                            custom_data[key] = val_str in ('evet', 'yes', 'true', '1', 'x', 'doğru')
                        elif field_type == 'number':
                            try:
                                custom_data[key] = float(cell_value) if cell_value else None
                            except (ValueError, TypeError):
                                row_errors.append(f"'{headers[col_idx]}' sayı olmalı")
                        elif field_type == 'date':
                            if hasattr(cell_value, 'strftime'):
                                custom_data[key] = cell_value.strftime('%Y-%m-%d')
                            else:
                                custom_data[key] = str(cell_value) if cell_value else None
                        else:
                            custom_data[key] = str(cell_value).strip() if cell_value else None
                    else:
                        custom_data[key] = None
                
                # Check required fields
                for req_key in required_fields:
                    if not custom_data.get(req_key):
                        # Find label for this key
                        label = next((f.get('label') for f in schema if f.get('key') == req_key), req_key)
                        row_errors.append(f"'{label}' zorunlu alan")
                
                if row_errors:
                    results["errors"].append({"row": row_idx, "errors": row_errors})
                else:
                    # Create asset
                    try:
                        Asset.objects.create(
                            asset_type=asset_type,
                            tenant=tenant,
                            custom_data=custom_data
                        )
                        results["success"] += 1
                    except Exception as e:
                        results["errors"].append({"row": row_idx, "errors": [str(e)]})
            
            return Response({
                "message": f"{results['success']} varlık başarıyla oluşturuldu.",
                "success_count": results["success"],
                "error_count": len(results["errors"]),
                "errors": results["errors"][:20]  # Limit error details
            })
            
        except Exception as e:
            return Response({"detail": f"Dosya işlenirken hata: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def preview_import(self, request):
        """
        Parse an Excel file and return its headers + first 5 data rows.
        Also detects whether the file is an app-generated template by looking for
        a hidden `_mapping` sheet. When detected, the caller can skip the
        mapping step and run the existing `bulk_import` flow.

        Multipart fields:
          file        — xlsx/xls
          header_row  — 1-based int (default 1; ignored when a template is detected)
        """
        file = request.FILES.get('file')
        if not file:
            return Response({"detail": "Dosya gerekli."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            header_row = int(request.data.get('header_row', 1))
        except (ValueError, TypeError):
            header_row = 1

        try:
            # Not read_only — we need access to sheet_state to detect hidden _mapping
            wb = openpyxl.load_workbook(file, data_only=True)

            # ── Template detection ──────────────────────────────────────────
            detected_template = None
            if '_mapping' in wb.sheetnames:
                mapping_ws = wb['_mapping']
                try:
                    raw_at_id = mapping_ws.cell(row=2, column=1).value
                    at_id = str(raw_at_id).strip() if raw_at_id else ''
                    if at_id:
                        try:
                            at = AssetType.objects.get(id=at_id)
                            detected_template = {
                                "asset_type_id": str(at.id),
                                "asset_type_name": at.name,
                            }
                        except (AssetType.DoesNotExist, ValueError, Exception):
                            detected_template = None
                except Exception:
                    detected_template = None

            # Templates always have headers on row 1 and data from row 4.
            if detected_template:
                header_row = 1

            ws = wb.active

            # Read headers
            headers = []
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=header_row, column=col_idx).value
                headers.append(str(val).strip() if val is not None else f"Sütun {col_idx}")

            # Trim trailing empty headers
            while headers and headers[-1].startswith("Sütun "):
                try:
                    int(headers[-1].split()[-1])
                    last_col = len(headers)
                    if ws.cell(row=header_row, column=last_col).value is None:
                        headers.pop()
                    else:
                        break
                except ValueError:
                    break

            # For templates, data starts at row 4 (header + info + example rows above).
            data_start = 4 if detected_template else header_row + 1

            # Return up to 50 preview rows so the UI can paginate them client-side.
            PREVIEW_ROW_LIMIT = 50
            preview_rows = []
            total_data_rows = 0
            for row_idx in range(data_start, ws.max_row + 1):
                row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, len(headers) + 1)]
                if all(v is None or str(v).strip() == '' for v in row_values):
                    continue
                total_data_rows += 1
                if len(preview_rows) < PREVIEW_ROW_LIMIT:
                    preview_rows.append([str(v) if v is not None else '' for v in row_values])

            wb.close()

            return Response({
                "headers": headers,
                "preview_rows": preview_rows,
                "total_rows": total_data_rows,
                "detected_template": detected_template,
            })

        except Exception as e:
            return Response({"detail": f"Dosya okunamadı: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def raw_import(self, request):
        """
        Import assets from an arbitrary Excel file using a user-supplied column mapping.

        Multipart fields:
          file         — xlsx/xls
          asset_type   — UUID
          tenant       — int (required for superuser)
          mapping      — JSON: { "Excel Sütun Adı": "schema_field_key" | null, ... }
                         null = ignore that column
          header_row   — int (default 1)
          skip_rows    — JSON array of first-cell string values to skip, e.g. ["TOPLAM"]
        """
        file = request.FILES.get('file')
        asset_type_id = request.data.get('asset_type')
        tenant_id = request.data.get('tenant')
        mapping_raw = request.data.get('mapping', '{}')
        skip_rows_raw = request.data.get('skip_rows', '[]')

        if not file:
            return Response({"detail": "Dosya gerekli."}, status=status.HTTP_400_BAD_REQUEST)
        if not asset_type_id:
            return Response({"detail": "Varlık tipi gerekli."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            mapping = json.loads(mapping_raw) if isinstance(mapping_raw, str) else mapping_raw
        except (json.JSONDecodeError, TypeError):
            return Response({"detail": "Geçersiz mapping JSON."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            skip_rows = json.loads(skip_rows_raw) if isinstance(skip_rows_raw, str) else skip_rows_raw
            skip_rows = [str(s).strip().lower() for s in skip_rows if s]
        except (json.JSONDecodeError, TypeError):
            skip_rows = []

        try:
            header_row = int(request.data.get('header_row', 1))
        except (ValueError, TypeError):
            header_row = 1

        try:
            asset_type = AssetType.objects.get(id=asset_type_id)
        except AssetType.DoesNotExist:
            return Response({"detail": "Varlık tipi bulunamadı."}, status=status.HTTP_404_NOT_FOUND)

        # Determine tenant
        user = request.user
        if user.is_superuser or user.is_platform_admin:
            if tenant_id:
                from apps.tenants.models import Tenant
                try:
                    tenant = Tenant.objects.get(id=tenant_id)
                except Tenant.DoesNotExist:
                    return Response({"detail": "Tersane bulunamadı."}, status=status.HTTP_404_NOT_FOUND)
            else:
                return Response({"detail": "Superuser için tersane seçimi gerekli."}, status=status.HTTP_400_BAD_REQUEST)
        else:
            tenant = user.tenant
            if not tenant:
                return Response({"detail": "Kullanıcının tersanesi yok."}, status=status.HTTP_400_BAD_REQUEST)

        schema = asset_type.schema or []
        field_types = {f['key']: f.get('type', 'text') for f in schema}

        # Find the unique key field (if any)
        unique_key_field = next((f['key'] for f in schema if f.get('is_unique_key')), None)

        try:
            wb = openpyxl.load_workbook(file, data_only=True, read_only=True)
            ws = wb.active

            # Read headers from the configured header row
            excel_headers = []
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=header_row, column=col_idx).value
                excel_headers.append(str(val).strip() if val is not None else '')

            # Build ordered list of schema field keys per column (None = ignore)
            col_keys = [mapping.get(h) for h in excel_headers]

            results = {"created": 0, "updated": 0, "skipped": 0, "errors": []}

            for row_idx in range(header_row + 1, ws.max_row + 1):
                row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, len(excel_headers) + 1)]

                # Skip fully empty rows
                if all(v is None or str(v).strip() == '' for v in row_values):
                    continue

                # Skip rows matching skip_rows patterns (check first non-None cell)
                first_val = str(row_values[0]).strip().lower() if row_values[0] is not None else ''
                if first_val in skip_rows:
                    results["skipped"] += 1
                    continue

                custom_data = {}
                row_errors = []

                for col_idx, key in enumerate(col_keys):
                    if not key:
                        continue
                    cell_value = row_values[col_idx] if col_idx < len(row_values) else None
                    field_type = field_types.get(key, 'text')

                    if cell_value is not None and str(cell_value).strip() != '':
                        if field_type == 'checkbox':
                            val_str = str(cell_value).lower().strip()
                            custom_data[key] = val_str in ('evet', 'yes', 'true', '1', 'x', 'doğru')
                        elif field_type == 'number':
                            try:
                                custom_data[key] = float(cell_value)
                            except (ValueError, TypeError):
                                row_errors.append(f"'{excel_headers[col_idx]}' sayı olmalı")
                        elif field_type == 'date':
                            if hasattr(cell_value, 'strftime'):
                                custom_data[key] = cell_value.strftime('%Y-%m-%d')
                            else:
                                custom_data[key] = str(cell_value).strip()
                        else:
                            custom_data[key] = str(cell_value).strip()
                    else:
                        custom_data[key] = None

                if row_errors:
                    results["errors"].append({"row": row_idx, "errors": row_errors})
                    continue

                try:
                    if unique_key_field and custom_data.get(unique_key_field):
                        unique_val = custom_data[unique_key_field]
                        lookup = {f"custom_data__{unique_key_field}": unique_val}
                        existing = Asset.objects.filter(
                            tenant=tenant,
                            asset_type=asset_type,
                            **lookup
                        ).first()
                        if existing:
                            merged = {**existing.custom_data, **{k: v for k, v in custom_data.items() if v is not None}}
                            existing.custom_data = merged
                            existing.save(update_fields=['custom_data', 'updated_at'])
                            results["updated"] += 1
                        else:
                            Asset.objects.create(
                                asset_type=asset_type,
                                tenant=tenant,
                                custom_data=custom_data,
                                created_by=user,
                            )
                            results["created"] += 1
                    else:
                        Asset.objects.create(
                            asset_type=asset_type,
                            tenant=tenant,
                            custom_data=custom_data,
                            created_by=user,
                        )
                        results["created"] += 1
                except Exception as e:
                    results["errors"].append({"row": row_idx, "errors": [str(e)]})

            wb.close()

            total = results["created"] + results["updated"]
            msg = f"{total} satır işlendi: {results['created']} oluşturuldu, {results['updated']} güncellendi."
            if results["skipped"]:
                msg += f" {results['skipped']} satır atlandı."

            return Response({
                "message": msg,
                "created_count": results["created"],
                "updated_count": results["updated"],
                "skipped_count": results["skipped"],
                "error_count": len(results["errors"]),
                "errors": results["errors"][:20],
            })

        except Exception as e:
            return Response({"detail": f"Dosya işlenirken hata: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def export(self, request):
        """
        Export assets to Excel file.
        """
        asset_type_id = request.query_params.get('asset_type')
        tenant_id = request.query_params.get('tenant')
        
        if not asset_type_id:
            return Response({"detail": "Varlık tipi gerekli."}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            asset_type = AssetType.objects.get(id=asset_type_id)
        except AssetType.DoesNotExist:
            return Response({"detail": "Varlık tipi bulunamadı."}, status=status.HTTP_404_NOT_FOUND)
        
        # Get assets based on user permissions
        queryset = self.get_queryset().filter(asset_type=asset_type)
        
        # Apply tenant filter if provided (for superusers)
        if tenant_id and (request.user.is_superuser or request.user.is_platform_admin):
            queryset = queryset.filter(tenant_id=tenant_id)
        
        schema = asset_type.schema or []
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = asset_type.name[:31]
        
        # Header styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Build headers
        headers = ['ID', 'Tersane', 'Oluşturulma Tarihi']
        for field in schema:
            headers.append(field.get('label', field.get('key', 'Alan')))
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 5, 15)
        
        # Write data
        for row_idx, asset in enumerate(queryset, 2):
            ws.cell(row=row_idx, column=1, value=str(asset.id)[:8])
            ws.cell(row=row_idx, column=2, value=asset.tenant.name if asset.tenant else '-')
            ws.cell(row=row_idx, column=3, value=asset.created_at.strftime('%Y-%m-%d %H:%M'))
            
            for col_idx, field in enumerate(schema, 4):
                key = field.get('key', '')
                value = asset.custom_data.get(key, '')
                
                if field.get('type') == 'checkbox':
                    value = 'Evet' if value else 'Hayır'
                
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Freeze header
        ws.freeze_panes = 'A2'
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"{asset_type.name.replace(' ', '_')}_listesi.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
